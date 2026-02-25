"""docker_scraper.py – YouTube live chat logger (Docker edition).

Drives a visible Firefox window (shown via Xvfb/noVNC so you can watch
and control it remotely) and monitors YouTube live chat by injecting a
JavaScript MutationObserver into the live-chat iframe.

Output: a timestamped XLSX file in LOGS_DIR with three sheets:

    Chat          – every message and moderation event
    VDS           – banned users only (mirrored from Chat)
    Livestream URL – the resolved stream URL captured at startup

Each sheet has five columns with a light-blue, bold, uppercase header row:

    TIME        Timestamp the message was captured (local time)
    USER        Author display name
    MESSAGE     Message text
    STATUS      blank | Deleted by user | Deleted by mod |
                Hidden | Timeout – X min | Banned
    MOD ACTION  Name of the moderator who performed the action (if known)

Delayed mod actions (e.g. a mod deletes a message 15 s after it was sent)
are handled by updating the original row in-place using an in-memory index
of {message_id → row_number}.

Environment variables
---------------------
YOUTUBE_CHANNEL_URL   Channel URL; the scraper navigates to URL/live to find
                      the current livestream.
                      Default: https://www.youtube.com
LOGS_DIR              Directory for the output XLSX.
                      Default: /app/Logs
POLL_INTERVAL         Seconds between DOM reads.  Default: 2
RETRY_INTERVAL        Seconds to wait before retrying when no livestream is
                      active.  Default: 60
"""

import logging
import os
import sys
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import JavascriptException, WebDriverException
from selenium.webdriver.firefox.options import Options

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger("docker_scraper")

# ── Configuration ─────────────────────────────────────────────────────────────
CHANNEL_URL    = os.getenv("YOUTUBE_CHANNEL_URL", "https://www.youtube.com")
LOGS_DIR       = os.getenv("LOGS_DIR", "/app/Logs")
POLL_INTERVAL  = float(os.getenv("POLL_INTERVAL", "2"))
RETRY_INTERVAL = int(os.getenv("RETRY_INTERVAL", "60"))

FIREFOX_PROFILE_DIR = "/app/firefox-profile"

# ── XLSX styling ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6",
                          fill_type="solid")
HEADER_FONT = Font(bold=True)
HEADERS     = ["TIME", "USER", "MESSAGE", "STATUS", "MOD ACTION"]

# ── JavaScript injected into the YouTube live-chat iframe ─────────────────────
#
# Installs a MutationObserver that pushes every new/removed message and every
# moderation system message into window._ytChatLog.  Python reads this queue
# periodically and resets it.
#
_JS_SETUP = r"""
if (!window._ytChatMonitorInstalled) {
    window._ytChatLog = [];

    function _isoNow() { return new Date().toISOString(); }

    /* Parse system messages shown to the channel owner (bans / timeouts). */
    function _parseSystemMsg(el) {
        var textEl = el.querySelector('#text, yt-formatted-string#text, .message');
        var text   = textEl ? textEl.textContent.trim() : el.textContent.trim();
        if (!text) return null;

        /* Timeout: "Username was timed out for X minutes" */
        var tmo = text.match(/^(.+?)\s+was timed out for\s+(\d+)\s*(\w+)/i);
        if (tmo) {
            return { name: tmo[1].trim(),
                     status: 'Timeout \u2013 ' + tmo[2] + ' ' + tmo[3],
                     modAction: '' };
        }
        /* Ban: "Username has been removed" / "Username was banned" */
        var ban = text.match(/^(.+?)\s+(has been removed|was banned)/i);
        if (ban) {
            return { name: ban[1].trim(), status: 'Banned', modAction: '' };
        }
        return null;
    }

    function _captureEl(el, isRemoved) {
        var tag = el.tagName ? el.tagName.toLowerCase() : '';

        /* Moderation system messages (timeouts, bans). */
        if (tag === 'yt-live-chat-system-message-renderer') {
            var parsed = _parseSystemMsg(el);
            if (parsed) {
                window._ytChatLog.push({
                    type:      'system',
                    time:      _isoNow(),
                    id:        '',
                    name:      parsed.name,
                    message:   '',
                    status:    parsed.status,
                    modAction: parsed.modAction
                });
            }
            return;
        }

        /* Regular and Super Chat messages. */
        var isMsg = tag === 'yt-live-chat-text-message-renderer' ||
                    tag === 'yt-live-chat-paid-message-renderer';
        if (!isMsg) return;

        var id     = el.id || el.getAttribute('data-id') || '';
        var nameEl = el.querySelector('#author-name');
        var msgEl  = el.querySelector('#message');
        var name   = nameEl ? nameEl.textContent.trim() : '';
        var msg    = msgEl  ? msgEl.textContent.trim()  : '';
        if (!name && !msg) return;

        window._ytChatLog.push({
            type:      isRemoved ? 'deletion' : 'message',
            time:      _isoNow(),
            id:        id,
            name:      name,
            message:   msg,
            /* Removed messages are assumed deleted by a moderator.
               YouTube does not expose the actor in the DOM for plain
               viewers; channel-owner sessions see system messages for
               bans/timeouts but not always for individual deletions. */
            status:    isRemoved ? 'Deleted by mod' : '',
            modAction: ''
        });
    }

    var container = document.querySelector(
        '#items.yt-live-chat-item-list-renderer, ' +
        'yt-live-chat-item-list-renderer #items'
    );
    if (container) {
        /* Capture messages already visible when the observer is installed. */
        container.querySelectorAll(
            'yt-live-chat-text-message-renderer, ' +
            'yt-live-chat-paid-message-renderer, ' +
            'yt-live-chat-system-message-renderer'
        ).forEach(function(el) { _captureEl(el, false); });

        new MutationObserver(function(muts) {
            muts.forEach(function(m) {
                m.addedNodes.forEach(function(n) {
                    if (n.nodeType === 1) _captureEl(n, false);
                });
                m.removedNodes.forEach(function(n) {
                    if (n.nodeType === 1) _captureEl(n, true);
                });
            });
        }).observe(container, { childList: true });

        window._ytChatMonitorInstalled = true;
    }
}
return !!window._ytChatMonitorInstalled;
"""

# Read accumulated events and reset the queue.
_JS_READ    = "var e=window._ytChatLog||[];window._ytChatLog=[];return e;"
# Check whether the observer is still alive (reset on page navigation).
_JS_IS_ALIVE = "return !!window._ytChatMonitorInstalled;"


# ── XLSX helpers ──────────────────────────────────────────────────────────────

def _write_header(ws) -> None:
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _new_workbook(stream_url: str = "") -> openpyxl.Workbook:
    """Create a fresh workbook with the three required sheets."""
    wb = openpyxl.Workbook()

    ws_chat       = wb.active
    ws_chat.title = "Chat"
    _write_header(ws_chat)

    ws_vds       = wb.create_sheet("VDS")
    _write_header(ws_vds)

    ws_url = wb.create_sheet("Livestream URL")
    hdr    = ws_url.cell(row=1, column=1, value="LIVESTREAM URL")
    hdr.fill = HEADER_FILL
    hdr.font = HEADER_FONT
    if stream_url:
        ws_url.cell(row=2, column=1, value=stream_url)

    return wb


def _append_row(ws, values: list) -> int:
    """Append a row and return its 1-based row number."""
    ws.append(values)
    return ws.max_row


def _update_row(ws, row: int, status: str, mod_action: str) -> None:
    ws.cell(row=row, column=4, value=status)
    ws.cell(row=row, column=5, value=mod_action)


# ── Firefox / Selenium setup ──────────────────────────────────────────────────

def _make_driver() -> webdriver.Firefox:
    os.makedirs(FIREFOX_PROFILE_DIR, exist_ok=True)
    opts        = Options()
    opts.profile = webdriver.FirefoxProfile(FIREFOX_PROFILE_DIR)
    return webdriver.Firefox(options=opts)


# ── Chat iframe helpers ───────────────────────────────────────────────────────

def _switch_to_chat_frame(driver) -> bool:
    """Switch Selenium context into the live-chat iframe when present."""
    driver.switch_to.default_content()
    try:
        frames = driver.find_elements("css selector", "iframe#chatframe")
        if frames:
            driver.switch_to.frame(frames[0])
            return True
    except Exception:
        pass
    return False


def _iso_to_local(iso: str) -> str:
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.astimezone().strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ── Main loop ─────────────────────────────────────────────────────────────────

def main() -> None:
    os.makedirs(LOGS_DIR, exist_ok=True)

    log.info("Starting Firefox (profile: %s)…", FIREFOX_PROFILE_DIR)
    driver = _make_driver()
    driver.maximize_window()

    # Navigate to the channel's /live URL — YouTube redirects to the active
    # stream automatically, or shows an offline page if none is live.
    live_url = CHANNEL_URL.rstrip("/") + "/live"
    log.info("Navigating to %s", live_url)

    stream_url = ""
    while not stream_url:
        driver.get(live_url)
        time.sleep(5)
        current = driver.current_url
        if "watch?v=" in current:
            stream_url = current
            log.info("Livestream found: %s", stream_url)
        else:
            log.info("No active livestream. Retrying in %d s…", RETRY_INTERVAL)
            time.sleep(RETRY_INTERVAL)

    # Create the timestamped XLSX
    ts        = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    xlsx_path = os.path.join(LOGS_DIR, f"chat-{ts}.xlsx")
    wb        = _new_workbook(stream_url)
    wb.save(xlsx_path)
    log.info("XLSX: %s", xlsx_path)

    ws_chat = wb["Chat"]
    ws_vds  = wb["VDS"]

    # {message_id: row_number_in_Chat_sheet} for in-place updates
    msg_row: dict = {}
    monitor_installed = False
    last_save         = time.time()
    dirty             = False

    log.info("Monitoring live chat… (press Ctrl+C to stop)")

    while True:
        try:
            _switch_to_chat_frame(driver)

            # (Re-)install the observer when needed
            if not monitor_installed:
                try:
                    monitor_installed = bool(driver.execute_script(_JS_SETUP))
                    if monitor_installed:
                        log.info("MutationObserver active — chat logging started.")
                except JavascriptException:
                    pass

            if monitor_installed:
                try:
                    events = driver.execute_script(_JS_READ) or []
                except JavascriptException:
                    events = []
                    monitor_installed = False

                for ev in events:
                    ts_str     = _iso_to_local(ev.get("time", ""))
                    ev_type    = ev.get("type", "message")
                    ev_id      = ev.get("id", "")
                    name       = ev.get("name", "")
                    message    = ev.get("message", "")
                    status     = ev.get("status", "")
                    mod_action = ev.get("modAction", "")

                    if ev_type == "message":
                        if ev_id and ev_id in msg_row:
                            continue  # already captured
                        row = _append_row(ws_chat,
                                          [ts_str, name, message, status, mod_action])
                        if ev_id:
                            msg_row[ev_id] = row
                        log.info("[MSG]  %s: %s", name, message)
                        dirty = True

                    elif ev_type == "deletion":
                        if ev_id and ev_id in msg_row:
                            # Update existing row in-place (handles delayed actions)
                            _update_row(ws_chat, msg_row[ev_id], status, mod_action)
                            log.info("[DEL]  updated row %d — %s: %s",
                                     msg_row[ev_id], name, message)
                        else:
                            # Message arrived before the observer was installed
                            row = _append_row(ws_chat,
                                              [ts_str, name, message, status, mod_action])
                            if ev_id:
                                msg_row[ev_id] = row
                            log.info("[DEL]  %s: %s", name, message)
                        dirty = True

                    elif ev_type == "system":
                        row = _append_row(ws_chat,
                                          [ts_str, name, message, status, mod_action])
                        if ev_id:
                            msg_row[ev_id] = row
                        # Mirror bans to the VDS sheet
                        if status == "Banned":
                            _append_row(ws_vds,
                                        [ts_str, name, message, status, mod_action])
                        log.info("[SYS]  %s → %s", name, status)
                        dirty = True

                # Check whether the observer is still alive after navigation
                try:
                    if not bool(driver.execute_script(_JS_IS_ALIVE)):
                        monitor_installed = False
                except Exception:
                    monitor_installed = False

            # Auto-save every 30 seconds when there is new data
            if dirty and (time.time() - last_save >= 30):
                wb.save(xlsx_path)
                last_save = time.time()
                dirty     = False
                log.debug("XLSX saved.")

        except WebDriverException as exc:
            log.warning("WebDriver error (browser navigating?): %s", exc)
            monitor_installed = False
            time.sleep(3)
        except KeyboardInterrupt:
            log.info("Interrupted — saving and exiting.")
            wb.save(xlsx_path)
            driver.quit()
            return
        except Exception as exc:
            log.error("Unexpected error: %s", exc)

        time.sleep(POLL_INTERVAL)


if __name__ == "__main__":
    main()
